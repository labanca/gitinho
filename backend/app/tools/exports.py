"""Excel exports — return an export_id the frontend can download."""

from __future__ import annotations

import io
from datetime import datetime, timedelta, timezone
from typing import Any
from uuid import UUID, uuid4

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.models import Export
from app.tools._base import ToolMode, registry
from app.tools._context import ToolContext
from app.tools.activity import org_users_activity_report
from app.tools.repos import _all_repos

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
EXPORT_TTL = timedelta(days=7)


def _build_workbook(title: str, rows: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    if not rows:
        ws.append(["(empty)"])
    else:
        headers = list(rows[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in rows:
            ws.append([_xlsx_cell(row.get(h)) for h in headers])
        for col_idx, h in enumerate(headers, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(
                12, min(40, len(str(h)) + 4)
            )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xlsx_cell(v: Any) -> Any:
    if isinstance(v, list | dict):
        import json as _json

        return _json.dumps(v, ensure_ascii=False)
    return v


async def _persist_export(
    db: AsyncSession,
    *,
    user_id: UUID,
    chat_id: UUID | None,
    filename: str,
    payload: bytes,
) -> UUID:
    export = Export(
        id=uuid4(),
        user_id=user_id,
        chat_id=chat_id,
        filename=filename,
        mime=XLSX_MIME,
        payload=payload,
        expires_at=datetime.now(timezone.utc) + EXPORT_TTL,
    )
    db.add(export)
    await db.commit()
    return export.id


@registry.register(mode=ToolMode.READ)
async def export_repos_xlsx(
    ctx: ToolContext,
    fields: list[str] | None = None,
    include_archived: bool = True,
) -> dict[str, Any]:
    """Export all repositories to an Excel file.

    `fields`: subset of [name, name_with_owner, description, is_private,
    is_archived, primary_language, branch_count, open_issues, open_prs,
    stargazers, forks, pushed_at, updated_at, created_at, topics, url].
    Defaults to a sensible set.
    """
    repos = await _all_repos(ctx)
    if not include_archived:
        repos = [r for r in repos if not r.is_archived]

    default_fields = [
        "name_with_owner",
        "is_private",
        "is_archived",
        "primary_language",
        "branch_count",
        "open_issues",
        "open_prs",
        "stargazers",
        "forks",
        "topics",
        "pushed_at",
        "updated_at",
        "created_at",
        "url",
    ]
    chosen = fields or default_fields
    rows = [{f: getattr(r, f, None) for f in chosen} for r in repos]

    payload = _build_workbook("Repositórios", rows)
    db = ctx.settings  # placeholder — real db injected in runner
    return {
        "kind": "export_xlsx",
        "filename": f"{ctx.org}-repos.xlsx",
        "payload_bytes": payload,
        "rows": len(rows),
    }


@registry.register(mode=ToolMode.READ)
async def export_users_activity_xlsx(
    ctx: ToolContext,
    since: str | None = None,
    until: str | None = None,
) -> dict[str, Any]:
    """Export the per-user activity report to an Excel file."""
    report = await org_users_activity_report(ctx, since=since, until=until)
    payload = _build_workbook("Atividade por usuário", report["rows"])
    return {
        "kind": "export_xlsx",
        "filename": f"{ctx.org}-user-activity.xlsx",
        "payload_bytes": payload,
        "rows": len(report["rows"]),
    }
